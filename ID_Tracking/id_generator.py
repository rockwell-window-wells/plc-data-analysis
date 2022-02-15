import pandas as pd
import qrcode
import os
from stat import S_IREAD, S_IRGRP, S_IROTH, S_IWUSR
from fpdf import FPDF
from openpyxl import load_workbook
import datetime as dt
import numpy as np
from PIL import Image, ImageDraw, ImageFont

##### ID card printing class and functions #####

# NOTE: The final version might need to work with image files, in which case
# these ID printing methods will need to change
def generate_idcard(templatefile, idnum, itemtype, qrcodepath, employee_name):
    # Change the background color based on the type of role or item
    prefix = str(idnum)[0:2]
    # Load template file (portrait CR80 with white background)
    template = Image.open(templatefile)
    W = template.size[0]
    # print("template width W is: {}".format(W))
    H = template.size[1]
    # print("template height H is: {}".format(H))
    bleed = 50

    redrgb = (255, 64, 64)
    bluergb = (97, 171, 255)
    whitergb = (255, 255, 255)
    equiprgb = (236, 232, 26)

    # Change the template from white to slightly gray for visibility
    template = template.convert('RGBA')
    data = np.array(template)
    red, green, blue, alpha = data.T
    white_areas = (red == 255) & (blue == 255) & (green == 255)
    # Red for supervisor
    if prefix == "10":
        data[..., :-1][white_areas.T] = redrgb
    # Blue for assistant
    elif  prefix == "11":
        data[..., :-1][white_areas.T] = bluergb
    else:
        data[..., :-1][white_areas.T] = equiprgb
        # data[..., :-1][white_areas.T] = whitergb
    template = Image.fromarray(data)
    template = template.convert('RGB')

    qr = Image.open(qrcodepath)
    # # Change the QR code background from white to slightly gray to match card
    # qr = qr.convert('RGBA')
    # data = np.array(qr)
    # red, green, blue, alpha = data.T
    # white_areas = (red == 255) & (blue == 255) & (green == 255)
    # # Red for supervisor
    # if prefix == "10":
    #     data[..., :-1][white_areas.T] = redrgb
    # # Blue for assistant
    # elif prefix == "11":
    #     data[..., :-1][white_areas.T] = bluergb
    # else:
    #     data[..., :-1][white_areas.T] = whitergb
    # qr = Image.fromarray(data)
    # qr = qr.convert('RGB')

    # Place QR code on the bottom of portrait card, resized to fit
    qrwidth = qr.size[0]
    qrheight = qr.size[1]
    # Determine the largest qr code that will fit in the bottom of the card
    qrfactor = 1
    if qrfactor*qrwidth > (W - 2*bleed):
        raise ValueError("QR code is too large for the chosen template")
    while qrfactor*qrwidth <= (W - 2*bleed):
        qrfactor += 1
        if qrfactor*qrwidth <= (W - 2*bleed):
            continue
        else:
            qrfactor -= 1
            break

    qr = qr.resize((qrfactor*qrwidth, qrfactor*qrheight))
    qrwidth = qr.size[0]
    qrheight = qr.size[1]
    border = (W - qrwidth)//2 - bleed
    left = bleed + border
    right = bleed + border + qrwidth
    bottom = H - border - bleed
    top = H - border - bleed - qrheight
    template.paste(qr, (left, top, right, bottom))

    fontsize = 75
    font = ImageFont.truetype("./assets/Roboto-Black.ttf", size=fontsize)
    draw = ImageDraw.Draw(template)
    if prefix != "11":
        draw.rectangle((bleed, bleed, W-bleed, H-bleed))
    if employee_name is not None:
        # print(employee_name)
        msg = employee_name
        w, h = draw.textsize(msg, font=font)
        while True:
            if w < (W - 2*bleed -10):
                break
            else:
                fontsize -= 5
                font = ImageFont.truetype("./assets/Roboto-Black.ttf", size=fontsize)
                w, h = draw.textsize(msg, font=font)
        draw.text(((W-w)/2, (H-W)/6 + bleed), employee_name, font=font, fill='black')
    fontsize = 75
    font = ImageFont.truetype("./assets/Roboto-Black.ttf", size=fontsize)
    msg = itemtype
    w, h = draw.textsize(msg, font=font)
    draw.text(((W-w)/2, (H-W)/2 + bleed), itemtype, font=font, fill='black')
    msg = str(idnum)
    w, h = draw.textsize(msg, font=font)
    draw.text(((W-w)/2, 5*(H-W)/6 + bleed), str(idnum), font=font, fill='black')
    return template


def print_IDcard_5digit(idnum, IDfilepath, QRfolder, IDcardfolder):
    """Take a given ID number (5-digit integer), find the corresponding name and
    QR code, and produce a printable ID card. Export it to the IDcardfolder.
    """
    # Get the necessary inputs to print_page function
    df = pd.read_excel(IDfilepath, None)
    sheetnames = df.keys()

    prefix = int(str(idnum)[0:2])
    foundID = False     # Flag for whether the chosen 5-digit ID exists
    # Find the appropriate ID sheet in the ID file
    for sheetname in sheetnames:
        iddata = df[sheetname]
        prefixcheck = int(str(iddata.loc[0, "ID"])[0:2])
        # Only look at the sheets with the right prefix. Skip those that aren't
        # the right one.
        if prefix != prefixcheck:
            continue
        else:
            # Pull out a copy of the row with the chosen ID
            idrow = iddata.loc[iddata["ID"] == idnum]
            idindex = idrow.index[0]
            itemtype = idrow.loc[idindex, "Type"]
            # Sanity check that idrow is only one row
            if len(idrow.index) != 1:
                raise Exception("[ERROR] There are 0 or more than 1 IDs that match")
            # elif pd.isna(idrow.loc[idindex, "Date"]):
            #     raise Exception("[ERROR] The chosen ID has not been assigned.")
            else:
                foundID = True
                if "Name" not in iddata.columns:
                    # print("no name found in data for this ID")
                    employee_name = None
                else:
                    employee_name = idrow.loc[idindex, "Name"]
                    # print("name found: {}".format(employee_name))

    if foundID is False:
        raise Exception("No ID found with number {}".format(idnum))

    qrcodename = str(idnum) + ".png"
    qrcodepath = QRfolder + '/' + qrcodename
    # Check if the QR code exists. If it doesn't, print the QR code.
    if not os.path.exists(qrcodepath):
        generate_qrcode(str(idnum), QRfolder)

    # Print the ID card and export it to the IDcardfolder with
    templatefile = "./assets/Portrait_white_ID.png"
    card = generate_idcard(templatefile, idnum, itemtype, qrcodepath, employee_name)
    IDcardfilename = str(idnum) + itemtype + ".png"
    IDcardpath = IDcardfolder + '/' + IDcardfilename
    card.save(IDcardpath)

    # pdf = IDCardPDF()
    # pdf.print_page(idnum, itemtype, qrcodepath, employee_name)
    # IDcardfilename = str(idnum) + itemtype + ".pdf"
    # IDcardpath = IDcardfolder + '/' + IDcardfilename
    # pdf.output(IDcardpath, 'F')

def print_IDcard_type_3digit(typestring, id_num, IDfilepath, QRfolder, IDcardfolder):
    """For cases where a card needs to be reprinted, this function takes a
    type string (i.e. "moldbacker" or "purple") and a 3-digit ID number, checks
    against the records in an Excel file, and if the desired ID exists, prints
    the corresponding QR code and ID card
    """
    typedict = {"purple": 30,
                "bag": 31,
                "pictureframe": 32}
    prefixes = []
    id_strings = []
    df = pd.read_excel(IDfilepath, None)
    sheetnames = df.keys()

    if typestring in typedict.keys():
        prefixes.append(typedict.get(typestring))
    elif typestring == "personnel":
        # Iterate through the sheets
        for sheetname in sheetnames:
            iddata = df[sheetname]
            print(iddata)
            # All personnel sheets have a Name column. Equipment sheets don't have
            # this column, so we use it to catch only the personnel relevant data.
            if "Name" in iddata:
                IDexample = str(iddata.loc[0,"ID"])
                prefixes.append(int(IDexample[0:2]))
    else:
        raise ValueError("[ERROR] typestring is not in approved list")

    # Combine the prefixes and IDs to get a complete list of IDs to look for.
    # This is probably going to be a list of length 1, but if it's for personnel
    # then it will be longer.
    for i in range(len(prefixes)):
        id_str = generate_id_string(prefixes[i], id_num)
        id_strings.append(id_str)

    id_ints = []
    for string in id_strings:
        id_ints.append(int(string))
        print("id_ints: {}".format(id_ints))

    # Check to verify that all of the ID numbers have been assigned
    for id in id_ints:
        print("\nChecking for ID {}".format(id))
        idgood = False
        for sheetname in sheetnames:
            iddata = df[sheetname]
            print(sheetname)
            if (id in iddata["ID"].values) and (not pd.isna(iddata.loc[(id_num-1),"Date"])):
                print("found ID in {}".format(sheetname))
                idgood = True
                if idgood is False:
                    # raise ValueError("ID {} has not been assigned yet".format(id))
                    print("[ERROR] ID {} has not been assigned yet".format(id))
                    # Remove the bad ID from id_ints
                    id_ints.remove(id)

    # If all the ID numbers have been assigned, then print their QR codes
    list_QR(id_ints, QRfolder)
    # Print the ID cards once the QR cards are generated
    list_IDcard(id_ints, IDfilepath, QRfolder, IDcardfolder)

def print_all_employee_IDcards(IDfilepath, QRfolder, IDcardfolder):
    allnums = get_all_employee_nums(IDfilepath)
    indices = allnums.index

    # Load the dataframe with all ID data
    df = pd.read_excel(IDfilepath, None)
    sheetnames = df.keys()

    # Iterate through the sheets
    for sheetname in sheetnames:
        iddata = df[sheetname]

        # All personnel sheets have a Name column. Equipment sheets don't have
        # this column, so we use it to catch only the personnel relevant data.
        if "Name" in iddata:
            for index in indices:
                id_int = iddata.loc[index, "ID"]
                id_string = str(id_int)
                # print(id_string)
                generate_qrcode(id_string, QRfolder)
                print_IDcard_5digit(id_int, IDfilepath, QRfolder, IDcardfolder)

def list_IDcard(idlist, IDfilepath, QRfolder, IDcardfolder):
    """Take a list of ID numbers (full ID with prefixes, as integers) and print
    all of the corresponding ID cards as PDF documents. Only prints ID cards for
    ID numbers that exist in the ID Excel file.
    """
    for id in idlist:
        try:
            print_IDcard_5digit(id, IDfilepath, QRfolder, IDcardfolder)
        except Exception as e:
            print(e)
            continue

def print_all_ID_by_type(typestring, IDfilepath, QRfolder, IDcardfolder):
    """Choose a category of ID from a list of acceptable strings, and print all
    of ID cards for any assigned ID numbers in that category. Personnel option
    prints all IDs that are associated with personnel (prefixes 10 and 11 as of
    1/12/2022).
    """
    typedict = {"purple": 30,
                "bag": 31,
                "pictureframe": 32}
    prefixes = []
    id_strings = []
    df = pd.read_excel(IDfilepath, None)
    sheetnames = df.keys()

    # Get the corresponding prefixes for the IDtype of interest
    if typestring in typedict.keys():
        prefixes.append(typedict.get(typestring))
    elif typestring == "personnel":
        # Iterate through the sheets
        for sheetname in sheetnames:
            iddata = df[sheetname]
            # All personnel sheets have a Name column. Equipment sheets don't have
            # this column, so we use it to catch only the personnel relevant data.
            if "Name" in iddata:
                # print("{}:\n{}".format(sheetname,iddata))
                IDexample = str(iddata.loc[0,"ID"])
                prefixes.append(int(IDexample[0:2]))
    else:
        raise ValueError("[ERROR] typestring is not in approved list")

    for prefix in prefixes:
        # Find the appropriate ID sheet in the ID file
        for sheetname in sheetnames:
            iddata = df[sheetname]
            prefixcheck = int(str(iddata.loc[0, "ID"])[0:2])

            # Once we're on the right data sheet for the chosen equipment type...
            if prefixcheck == prefix:
                # Get list of ID numbers that have been assigned
                assigned = iddata["ID"].loc[pd.notna(iddata["Date"])].tolist()
                list_IDcard(assigned, IDfilepath, QRfolder, IDcardfolder)

def print_all_IDcards(IDfilepath, QRfolder, IDcardfolder):
    """Single function to update the QR codes and ID cards with all of the
    currently assigned ID numbers for both personnel and equipment.
    """
    typekeys = ["personnel", "purple", "bag", "pictureframe"]
    for key in typekeys:
        print_all_ID_by_type(key, IDfilepath, QRfolder, IDcardfolder)


def N_new_equip_ids(n, typestring, IDfilepath, QRfolder, IDcardfolder):
    """Generate the next n equipment ID numbers, QR codes, and ID cards for the
    chosen equipment type.
    """
    typedict = {"purple": 30,
                "bag": 31,
                "pictureframe": 32}
    prefix = None

    # id_strings = []
    df = pd.read_excel(IDfilepath, None)
    sheetnames = df.keys()

    if typestring in typedict.keys():
        prefix = typedict.get(typestring)
    else:
        raise ValueError("[ERROR] typestring is not in approved list")

    # Find the appropriate ID sheet in the ID file
    for sheetname in sheetnames:
        # print(sheetname)
        iddata = df[sheetname]
        prefixcheck = int(str(iddata.loc[0, "ID"])[0:2])
        # print(prefixcheck)

        # Once we're on the right data sheet for the chosen equipment type...
        if prefixcheck == prefix:
            # find the first ID that doesn't have an assigned date filled in
            i = 0
            while i < len(iddata):
                # If the date value for the current index is not empty, continue
                if not pd.isna(iddata.loc[i,"Date"]):
                    i += 1
                    continue
                else:
                    indexA = i
                    break

            # We are going to include the indices indexA and indexB in the list.
            # Generate that list below.
            indexB = indexA + n - 1
            indlist = []
            if n == 1:
                indlist.append(indexA)
            else:
                while indexA <= indexB:
                    indlist.append(indexA)
                    indexA += 1

            # Now get the ID numbers corresponding to the list
            idlist = []
            for ind in indlist:
                idlist.append(iddata.loc[ind, "ID"])
                iddata.loc[ind, "Date"] = dt.date.today()

            # print(idlist)
            # Print the QR codes of all the IDs on the list
            list_QR(idlist, QRfolder)
            list_IDcard(idlist, IDfilepath, QRfolder, IDcardfolder)

    # Save the updated data to the Excel ID file
    rewrite_whole_Excel_sheet(IDfilepath, df, sheetnames)


##### QR code generation methods #####
def generate_qrcode(id_string, path):
    img = qrcode.make(id_string)
    filename = id_string + ".png"
    filepath = path + "\\" + filename
    img.save(filepath)

def range_QR(typeprefix, idA, idB, path):
    """Generate a sequential series of QR codes.

    Args:
        typeprefix: A 2-digit integer number corresponding to the type code.
        idA: An integer number of up to 3 digits, which is the initial ID number
            to be generated.
        idB: An integer number of up to 3 digits, which is the final ID number
            to be generated. Must be greater than idA
    Returns:
        N/A
    """
    if idB > idA:
        for i in range(idA, idB+1):
            id_string = generate_id_string(typeprefix, i)
            generate_qrcode(id_string, path)
    else:
        raise ValueError("idB must be greater than idA")

def list_QR(idlist, QRfolder):
    """Take a list of ID numbers (full ID with prefixes, as integers) and print
    all of the corresponding qr codes.
    """
    for id in idlist:
        id_string = str(id)
        # Sanity check on length of id_string
        if len(id_string) != 5:
            raise ValueError("[ERROR] ID number {} has the wrong number of digits.".format(id))
        else:
            generate_qrcode(id_string, QRfolder)


############### THIS NEEDS A CATEGORY ####################
def generate_id_string(typeprefix, idnum):
    """ Take the ID number integer (without leading zeros) and add leading zeros
    as well as a type prefix number.
    """
    len_prefix = len(str(typeprefix))
    len_idnum = len(str(idnum))
    if len_prefix != 2 or len_idnum < 1 or len_idnum > 3:
        raise ValueError("Prefix or ID number have an invalid number of digits")
    else:
        n = 3   # The idnum should have three digits
        # Add any preceding zeros to get the idnum to 3 digits
        idnum_str = str(idnum).zfill(n)
        # Combine the prefix and the ID number
        id_string = str(typeprefix) + idnum_str

        return id_string


##### Methods for employees to choose or modify their ID number #####
def assign_employee_nums_from_sheet(IDinputfile, IDfilepath, QRfolder, IDcardfolder):
    """Take an Excel sheet with two columns: Name and 3 digit ID number. Assign
    all of the numbers to employees and generate their QR codes and ID cards.
    Perform checks and catch any numbers that are duplicates in the input file
    or are already taken in the ID data file.

    Args:
        IDinputfile: The path to the input Excel file with a name column and an
            ID column.
        IDfilepath: The filepath of the Excel file where
            all the ID numbers for personnel and equipment are stored.
    Returns:
        N/A
    """
    # Load the input workbook, assuming the data is on the first sheet
    df_input = pd.read_excel(IDinputfile)

    # Double check if all the ID numbers and names are unique
    names = df_input["Name"].unique()
    numbers = df_input["ID"].unique()
    names = [x for x in names if pd.isnull(x) == False and x != 'nan']
    numbers = [x for x in numbers if pd.isnull(x) == False and x != 'nan']

    n_names = len(names)
    print(names)
    print(n_names)
    n_numbers = len(numbers)
    print(numbers)
    print(n_numbers)
    if n_names != n_numbers:
        raise Exception("A name or ID is missing from the input Excel sheet")

    exceptions = []
    for i in range(len(df_input)):
        desired_number = str(df_input.loc[i, "ID"])
        employee_name = df_input.loc[i, "Name"]

        try:
            assign_employee_num(desired_number, employee_name, IDfilepath, QRfolder, IDcardfolder)
        except Exception as e:
            exceptions.append(str(e))
            continue

    # Print out a list of all the exceptions at the end
    exceptions_str = ""
    for e in exceptions:
        exceptions_str = exceptions_str + e + "\n"
    print(exceptions_str)

def assign_employee_num(desired_number, employee_name, IDfilepath, QRfolder, IDcardfolder):
    """Allow an employee to pick their own ID number (3 digits). Check against
    an Excel spreadsheet to see if the number is taken. If it's taken, show an
    error message. If it's not taken, then assign the employee's name to the
    number for both the Lead and Assistant roles. Add the date the number was
    assigned for record keeping.

    Args:
        desired_number: An integer of up to 3 digits.
        employee_name: A string of the employee's name.
        IDfilepath: The filepath (including file name) of the Excel file where
            all the ID numbers for personnel and equipment are stored.
    Returns:
        N/A
    """
    # Load the workbook with all sheets (that's what the None flag is for)
    # df is a dictionary of sheet names and dataframes of the sheets
    df = pd.read_excel(IDfilepath, None)
    sheetnames = df.keys()

    # Iterate through the sheets
    for sheetname in sheetnames:
        iddata = df[sheetname]

        # All personnel sheets have a Name column. Equipment sheets don't have
        # this column, so we use it to catch only the personnel relevant data.
        if "Name" in iddata:
            IDexample = str(iddata.loc[0,"ID"])
            prefix = IDexample[0:2]

            # Handle the incoming desired_number, whether it is an integer or string
            if type(desired_number) is int:
                numstr = str(desired_number)
                n = 3   # The idnum should have three digits
                # Add any preceding zeros to get the idnum to 3 digits
                idnum_str = numstr.zfill(n)
            elif type(desired_number) is str:
                if len(desired_number) == 3:
                    idnum_str = desired_number
                elif len(desired_number) < 3:
                    nzeros = 3 - len(desired_number)
                    idnum_str = nzeros*"0" + desired_number
                else:
                    raise Exception("Desired ID number is too long. Choose one with 3 or less characters")
            else:
                raise Exception("Desired ID number was not entered as an integer or string")

            # Combine the prefix and the ID number
            num = prefix + idnum_str
            num = int(num)

            # Get the rows in leads and assistants that correspond to the desired number
            numind = iddata.index[iddata["ID"] == num]
            numind = numind[0]

            if pd.isna(iddata.loc[numind, "Name"]):
                # Update the Name column
                iddata.loc[numind, "Name"] = employee_name
                # Update the Date column
                iddata.loc[numind, "Date"] = dt.date.today()
                df[sheetname] = iddata
                print("ID {} assigned to {}".format(num, employee_name))
                rewrite_whole_Excel_sheet(IDfilepath, df, sheetnames)
                print_IDcard_5digit(num, IDfilepath, QRfolder, IDcardfolder)
            else:
                raise Exception("[ERROR] ID number {} has already been assigned.".format(num))


def get_all_employee_nums(IDfilepath):
    """Returns dataframe with collapsed list of all 3-digit ID numbers currently
    in use for employees only.
    """
    leads = pd.read_excel(IDfilepath, sheet_name="Personnel-Lead")
    allnums = leads[~leads["Name"].isnull()]   # Dataframe of just the rows with names assigned to IDs
    allnums = allnums.drop(columns=["Type", "Date"])
    allnums["ID"] = allnums["ID"].apply(str)
    allnums["ID"] = allnums["ID"].str[2:]
    return allnums

def reassign_employee_num(employee_name, new_desired_num):
    """Take an employee name (assumed to be unique in the list of employee names
    and ID numbers) and reassign them to a new desired ID number. Clear the data
    for the old number so it becomes available.
    """
    pass

##### Convenience methods #####
def rewrite_whole_Excel_sheet(IDfilepath, df, sheetnames):
    """Convenience function for writing all new data to the whole Excel sheet.
    """
    with pd.ExcelWriter(IDfilepath, engine='xlsxwriter',
                        date_format="yyyy-mm-dd",
                        datetime_format="yyyy-mm-dd") as writer:
        for sheetname in sheetnames:
            df[sheetname].to_excel(writer, sheet_name=sheetname, index=False)

# def make_readwrite(filepath):
#     """More readable version of os package method for making a file read/write.
#     """
#     os.chmod(filepath, S_IWUSR|S_IREAD)
#
# def make_readonly(filepath):
#     """More readable version of os package method for making a file read-only.
#     """
#     os.chmod(filepath, S_IREAD|S_IRGRP|S_IROTH)


##### Main function #####
if __name__ == '__main__':
    print("id_generator is being run as the main function")
